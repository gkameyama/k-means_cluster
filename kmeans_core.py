import csv
import logging
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook

LOG_FILE = "kmeans_log.txt"
SUPPORTED_INPUT_EXTENSIONS = (".csv", ".xlsx")
SUPPORTED_OUTPUT_EXTENSIONS = (".csv", ".xlsx")
KMEANS_RANDOM_STATE = 42
KMEANS_N_INIT = 10
KMEANS_MAX_ITER = 300
KMEANS_TOL = 1e-4


def setup_logger(log_to_stdout: bool = True) -> None:
    handlers = [logging.FileHandler(LOG_FILE, encoding="utf-8")]
    if log_to_stdout:
        handlers.append(logging.StreamHandler(sys.stdout))

    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)
    root_logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    for handler in handlers:
        handler.setFormatter(formatter)
        root_logger.addHandler(handler)


def read_csv_table(path: str) -> tuple[list[str], list[list[object]]]:
    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            with open(path, "r", encoding=encoding, newline="") as file:
                rows = list(csv.reader(file))
            break
        except Exception:
            rows = None
            continue
    else:
        raise ValueError("CSV を読み込めません。文字コードを確認してください。")

    if not rows:
        return [], []
    return rows[0], rows[1:]


def read_xlsx_table(path: str) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        values = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not values:
        return [], []
    headers = ["" if value is None else str(value) for value in values[0]]
    rows = [list(row) for row in values[1:]]
    return headers, rows


def read_table(path: str) -> tuple[list[str], list[list[object]]]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return read_csv_table(path)
    if ext == ".xlsx":
        return read_xlsx_table(path)
    raise ValueError("対応している入力形式は .csv / .xlsx のみです。")


def write_result(result_rows: list[tuple[object, int]], output_file: str) -> None:
    out_ext = Path(output_file).suffix.lower()
    if out_ext == ".csv":
        with open(output_file, "w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(["sample_id", "cluster"])
            writer.writerows(result_rows)
        return

    if out_ext == ".xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "result"
        worksheet.append(["sample_id", "cluster"])
        for row in result_rows:
            worksheet.append(list(row))
        workbook.save(output_file)
        workbook.close()
        return

    raise ValueError("対応している出力形式は .csv / .xlsx のみです。")


def build_timestamped_output_path(output_file: str) -> str:
    path = Path(output_file)
    timestamp = datetime.now().strftime("%m%d%H%M")
    candidate = path.with_name(f"{path.stem}_{timestamp}{path.suffix}")

    if not candidate.exists():
        return str(candidate)

    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem}_{timestamp}_{counter:02d}{path.suffix}")
        if not candidate.exists():
            return str(candidate)
        counter += 1


def summarize_result(
    input_file: str,
    output_file: str,
    id_col: str,
    original_rows: int,
    processed_rows: int,
    missing_rows: int,
    non_numeric_rows: int,
    n_clusters: int,
) -> dict:
    return {
        "input_file": input_file,
        "output_file": output_file,
        "id_column": id_col,
        "original_rows": original_rows,
        "processed_rows": processed_rows,
        "missing_rows_removed": int(missing_rows),
        "non_numeric_rows_removed": int(non_numeric_rows),
        "cluster_count": n_clusters,
    }


def format_summary(summary: dict) -> str:
    return "\n".join(
        [
            "クラスタリングが完了しました。",
            f"入力ファイル: {summary['input_file']}",
            f"出力ファイル: {summary['output_file']}",
            f"ID 列: {summary['id_column']}",
            f"元データ件数: {summary['original_rows']}",
            f"分析対象件数: {summary['processed_rows']}",
            f"欠損で除外した件数: {summary['missing_rows_removed']}",
            f"非数値で除外した件数: {summary['non_numeric_rows_removed']}",
            f"クラスタ数: {summary['cluster_count']}",
        ]
    )


def is_missing(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def convert_to_float(value: object) -> float:
    if isinstance(value, str):
        value = value.strip()
    return float(value)


def standardize_features(features: np.ndarray) -> np.ndarray:
    data = np.array(features, dtype=np.float64, copy=True)
    means = data.mean(axis=0)
    scales = data.std(axis=0, ddof=0)
    scales[scales == 0.0] = 1.0
    return (data - means) / scales


def row_norms_squared(data: np.ndarray) -> np.ndarray:
    return np.einsum("ij,ij->i", data, data)


def compute_squared_distances(data: np.ndarray, centers: np.ndarray) -> np.ndarray:
    data_norms = row_norms_squared(data)[:, np.newaxis]
    center_norms = row_norms_squared(centers)[np.newaxis, :]
    distances = data_norms + center_norms - (2.0 * data @ centers.T)
    return np.maximum(distances, 0.0)


def stable_cumsum(values: np.ndarray) -> np.ndarray:
    return np.cumsum(values, dtype=np.float64)


def init_centroids_kmeans_plus_plus(data: np.ndarray, n_clusters: int, rng: np.random.RandomState) -> np.ndarray:
    n_samples, n_features = data.shape
    centers = np.empty((n_clusters, n_features), dtype=data.dtype)

    center_id = rng.choice(n_samples, p=np.full(n_samples, 1.0 / n_samples))
    centers[0] = data[center_id]

    closest_dist_sq = compute_squared_distances(data, centers[:1]).reshape(-1)
    current_potential = closest_dist_sq.sum()
    n_local_trials = 2 + int(np.log(n_clusters))

    for center_index in range(1, n_clusters):
        if not np.isfinite(current_potential) or current_potential <= 0.0:
            candidate_ids = np.array([rng.randint(n_samples)], dtype=np.int64)
        else:
            random_values = rng.uniform(size=n_local_trials) * current_potential
            cumulative_distances = stable_cumsum(closest_dist_sq)
            candidate_ids = np.searchsorted(cumulative_distances, random_values, side="left")
            candidate_ids = np.clip(candidate_ids, None, n_samples - 1)

        distance_to_candidates = compute_squared_distances(data, data[candidate_ids]).T
        np.minimum(closest_dist_sq, distance_to_candidates, out=distance_to_candidates)
        candidate_potentials = distance_to_candidates.sum(axis=1)
        best_candidate_position = int(np.argmin(candidate_potentials))
        best_candidate = int(candidate_ids[best_candidate_position])
        best_potential = float(candidate_potentials[best_candidate_position])
        best_distances = distance_to_candidates[best_candidate_position]

        centers[center_index] = data[best_candidate]
        closest_dist_sq = best_distances
        current_potential = best_potential

    return centers


def assign_labels(data: np.ndarray, centers: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    distances = compute_squared_distances(data, centers)
    labels = np.argmin(distances, axis=1)
    min_distances = distances[np.arange(data.shape[0]), labels]
    return labels, min_distances


def relocate_empty_clusters(
    data: np.ndarray,
    centers: np.ndarray,
    labels: np.ndarray,
    distances_to_assigned: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    counts = np.bincount(labels, minlength=centers.shape[0]).astype(np.int64)
    empty_clusters = np.flatnonzero(counts == 0)
    if empty_clusters.size == 0:
        return labels, counts

    farthest_indices = np.argsort(distances_to_assigned)[::-1]
    for empty_cluster, sample_index in zip(empty_clusters, farthest_indices):
        source_cluster = labels[sample_index]
        if counts[source_cluster] <= 1:
            continue
        labels[sample_index] = empty_cluster
        centers[empty_cluster] = data[sample_index]
        counts[empty_cluster] += 1
        counts[source_cluster] -= 1

    return labels, counts


def recompute_centers(data: np.ndarray, labels: np.ndarray, n_clusters: int) -> np.ndarray:
    centers = np.zeros((n_clusters, data.shape[1]), dtype=np.float64)
    for cluster_index in range(n_clusters):
        members = data[labels == cluster_index]
        if len(members) > 0:
            centers[cluster_index] = members.mean(axis=0)
    return centers


def run_single_kmeans(data: np.ndarray, n_clusters: int, rng: np.random.RandomState) -> tuple[np.ndarray, np.ndarray, float]:
    centers = init_centroids_kmeans_plus_plus(data, n_clusters, rng)

    for _ in range(KMEANS_MAX_ITER):
        labels, distances_to_assigned = assign_labels(data, centers)
        labels, counts = relocate_empty_clusters(data, centers, labels, distances_to_assigned)

        if np.any(counts == 0):
            counts = np.bincount(labels, minlength=n_clusters).astype(np.int64)

        new_centers = recompute_centers(data, labels, n_clusters)
        center_shift = np.sum((centers - new_centers) ** 2)
        centers = new_centers

        if center_shift <= KMEANS_TOL:
            break

    labels, distances_to_assigned = assign_labels(data, centers)
    inertia = float(distances_to_assigned.sum())
    return labels, centers, inertia


def is_same_clustering(labels_a: np.ndarray | None, labels_b: np.ndarray | None, n_clusters: int) -> bool:
    if labels_a is None or labels_b is None or labels_a.shape != labels_b.shape:
        return False

    mapping = np.full(n_clusters, -1, dtype=np.int64)
    for label_a, label_b in zip(labels_a, labels_b):
        mapped = mapping[label_a]
        if mapped == -1:
            mapping[label_a] = label_b
        elif mapped != label_b:
            return False
    return True


def run_numpy_kmeans(data: np.ndarray, n_clusters: int) -> np.ndarray:
    rng = np.random.RandomState(KMEANS_RANDOM_STATE)
    best_labels = None
    best_inertia = None

    for _ in range(KMEANS_N_INIT):
        labels, _, inertia = run_single_kmeans(data, n_clusters, rng)
        if best_inertia is None or (
            inertia < best_inertia and not is_same_clustering(labels, best_labels, n_clusters)
        ):
            best_labels = labels.copy()
            best_inertia = inertia

    return best_labels + 1


def run_kmeans(input_file: str, n_clusters: int, output_file: str) -> tuple[list[tuple[object, int]], dict]:
    headers, rows = read_table(input_file)
    if len(headers) < 2:
        raise ValueError("データは 2 列以上必要です。先頭列を ID、2 列目以降を特徴量にしてください。")

    id_col = headers[0]
    feature_cols = headers[1:]
    logging.info("入力ファイル: %s", input_file)
    logging.info("ID 列: %s", id_col)
    logging.info("使用する特徴量列数: %s", len(feature_cols))

    original_rows = len(rows)
    missing_rows = 0
    non_numeric_rows = 0
    valid_ids: list[object] = []
    valid_features: list[list[float]] = []

    for row in rows:
        padded_row = list(row[: len(headers)]) + [None] * max(0, len(headers) - len(row))
        feature_values = padded_row[1 : len(headers)]

        if any(is_missing(value) for value in feature_values):
            missing_rows += 1
            continue

        try:
            numeric_row = [convert_to_float(value) for value in feature_values]
        except (TypeError, ValueError):
            non_numeric_rows += 1
            continue

        valid_ids.append(padded_row[0])
        valid_features.append(numeric_row)

    if missing_rows:
        logging.warning("欠損値により除外した行数: %s", missing_rows)
    if non_numeric_rows:
        logging.warning("非数値データにより除外した行数: %s", non_numeric_rows)

    if not valid_features:
        raise ValueError("分析可能なデータがありません。欠損値や非数値データを確認してください。")
    if n_clusters < 1:
        raise ValueError("クラスタ数は 1 以上を指定してください。")
    if n_clusters > len(valid_features):
        raise ValueError("クラスタ数は分析対象件数以下にしてください。")

    scaled_features = standardize_features(np.asarray(valid_features, dtype=np.float64))
    labels = run_numpy_kmeans(scaled_features, n_clusters)

    actual_output_file = build_timestamped_output_path(output_file)
    result_rows = list(zip(valid_ids, labels.tolist()))
    write_result(result_rows, actual_output_file)

    summary = summarize_result(
        input_file=input_file,
        output_file=actual_output_file,
        id_col=str(id_col),
        original_rows=original_rows,
        processed_rows=len(valid_features),
        missing_rows=missing_rows,
        non_numeric_rows=non_numeric_rows,
        n_clusters=n_clusters,
    )

    logging.info("元データ件数: %s", original_rows)
    logging.info("分析対象件数: %s", len(valid_features))
    logging.info("出力ファイル: %s", actual_output_file)
    logging.info("処理完了")
    return result_rows, summary
